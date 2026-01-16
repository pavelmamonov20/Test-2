import os
import sys
from tkinter import *
from tkinter import ttk, messagebox, filedialog
from barcode import Code128
from barcode.writer import ImageWriter
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
import tempfile
import threading


class BarcodeGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Генератор Штрихкодов")
        self.root.geometry("600x400")
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        # Single barcode tab
        self.single_frame = Frame(self.notebook)
        self.notebook.add(self.single_frame, text="Один штрихкод")
        self.create_single_tab()
        
        # Range barcode tab
        self.range_frame = Frame(self.notebook)
        self.notebook.add(self.range_frame, text="Диапазон штрихкодов")
        self.create_range_tab()
    
    def create_single_tab(self):
        # Label for single barcode
        Label(self.single_frame, text="Введите номер для генерации штрихкода:", font=("Arial", 12)).pack(pady=20)
        
        # Entry for single barcode
        self.single_entry = Entry(self.single_frame, font=("Arial", 12), width=20)
        self.single_entry.pack(pady=10)
        self.single_entry.insert(0, "001")  # Default value
        
        # Generate button for single barcode
        Button(self.single_frame, text="Сгенерировать штрихкод", 
               command=self.generate_single_barcode, 
               font=("Arial", 12), bg="#4CAF50", fg="white").pack(pady=20)
    
    def create_range_tab(self):
        # Labels and entries for range
        Label(self.range_frame, text="Диапазон штрихкодов (формат: CCXXX):", font=("Arial", 12)).pack(pady=20)
        
        range_inputs_frame = Frame(self.range_frame)
        range_inputs_frame.pack(pady=10)
        
        Label(range_inputs_frame, text="От:", font=("Arial", 12)).grid(row=0, column=0, padx=5)
        self.start_entry = Entry(range_inputs_frame, font=("Arial", 12), width=10)
        self.start_entry.grid(row=0, column=1, padx=5)
        self.start_entry.insert(0, "1")  # Default value
        
        Label(range_inputs_frame, text="До:", font=("Arial", 12)).grid(row=0, column=2, padx=5)
        self.end_entry = Entry(range_inputs_frame, font=("Arial", 12), width=10)
        self.end_entry.grid(row=0, column=3, padx=5)
        self.end_entry.insert(0, "200")  # Default value
        
        # Generate button for range
        Button(self.range_frame, text="Сгенерировать диапазон штрихкодов", 
               command=self.generate_range_barcodes, 
               font=("Arial", 12), bg="#2196F3", fg="white").pack(pady=20)
    
    def validate_number(self, num_str, field_name):
        """Validate that the input is a positive integer"""
        try:
            num = int(num_str)
            if num <= 0:
                messagebox.showerror("Ошибка", f"{field_name} должен быть положительным числом")
                return None
            return num
        except ValueError:
            messagebox.showerror("Ошибка", f"{field_name} должен быть числом")
            return None
    
    def generate_single_barcode(self):
        """Generate a single barcode with the CC prefix"""
        number_str = self.single_entry.get().strip()
        
        if not number_str:
            messagebox.showerror("Ошибка", "Введите номер штрихкода")
            return
        
        # Validate the input
        if not number_str.isdigit():
            messagebox.showerror("Ошибка", "Номер должен содержать только цифры")
            return
        
        # Format the number with leading zeros to ensure 3 digits
        formatted_num = f"{int(number_str):03d}"
        barcode_value = f"CC{formatted_num}"
        
        # Ask user for save location
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Сохранить штрихкод как",
            initialfile=f"single_barcode_{barcode_value}.xlsx"
        )
        
        if not filename:
            return  # User cancelled the dialog
        
        # Start generation in a separate thread to keep UI responsive
        threading.Thread(target=self._generate_single_excel, args=(barcode_value, filename), daemon=True).start()
    
    def _generate_single_excel(self, barcode_value, filename):
        """Generate Excel file with a single barcode"""
        try:
            # Show progress
            self.root.config(cursor="wait")
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Штрихкод"
            
            # Headers
            ws['A1'] = 'Номер'
            ws['B1'] = 'Штрихкод'
            
            # Column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 30
            
            # Row height
            ws.row_dimensions[2].height = 85.04  # ~30mm in points
            
            # Generate temporary barcode image
            temp_dir = tempfile.mkdtemp()
            try:
                barcode_filename = os.path.join(temp_dir, 'single_barcode')
                
                # Create barcode
                code128 = Code128(barcode_value, writer=ImageWriter())
                code128.save(barcode_filename)
                
                barcode_path = barcode_filename + '.png'
                
                if not os.path.exists(barcode_path):
                    raise FileNotFoundError(f"Barcode image was not created: {barcode_path}")
                
                # Add data to Excel
                ws['A2'] = barcode_value
                ws['B2'] = barcode_value
                
                # Add image to Excel
                excel_img = ExcelImage(barcode_path)
                
                # Scale image to fit cell
                target_width = 200  # pixels
                target_height = 100  # pixels
                excel_img.width = target_width
                excel_img.height = target_height
                
                ws.add_image(excel_img, 'B2')
                
                # Center align the image cell
                cell = ws['B2']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Save the workbook
                wb.save(filename)
                
                messagebox.showinfo("Успех", f"Штрихкод {barcode_value} успешно сохранен в {filename}")
                
            finally:
                # Clean up temp directory
                import shutil
                shutil.rmtree(temp_dir)
        
        except Exception as e:
            messagebox.showerror("Ошибка", f"Произошла ошибка при генерации файла:\n{str(e)}")
        finally:
            self.root.config(cursor="")
    
    def generate_range_barcodes(self):
        """Generate a range of barcodes with CC prefix"""
        start_str = self.start_entry.get().strip()
        end_str = self.end_entry.get().strip()
        
        if not start_str or not end_str:
            messagebox.showerror("Ошибка", "Введите оба числа для диапазона")
            return
        
        start_num = self.validate_number(start_str, "Начальное число")
        end_num = self.validate_number(end_str, "Конечное число")
        
        if start_num is None or end_num is None:
            return
        
        if start_num > end_num:
            messagebox.showerror("Ошибка", "Начальное число не может быть больше конечного")
            return
        
        # Ask user for save location
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Сохранить диапазон штрихкодов как",
            initialfile=f"range_barcodes_CC{start_num:03d}_to_CC{end_num:03d}.xlsx"
        )
        
        if not filename:
            return  # User cancelled the dialog
        
        # Start generation in a separate thread to keep UI responsive
        threading.Thread(target=self._generate_range_excel, args=(start_num, end_num, filename), daemon=True).start()
    
    def _generate_range_excel(self, start_num, end_num, filename):
        """Generate Excel file with range of barcodes"""
        try:
            # Show progress
            self.root.config(cursor="wait")
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Штрихкоды"
            
            # Headers
            ws['A1'] = 'Номер'
            ws['B1'] = 'Штрихкод'
            
            # Column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 30
            
            def _excel_col_width_to_pixels(width_chars):
                return int(width_chars * 7 + 5)
            
            col_width = ws.column_dimensions['B'].width
            if col_width is None:
                col_width = ws.sheet_format.defaultColWidth
            if col_width is None:
                col_width = 8.43
            cell_width_pixels = int(_excel_col_width_to_pixels(col_width))
            
            # Temporary directory for barcode images
            temp_dir = tempfile.mkdtemp()
            
            try:
                for i in range(start_num, end_num + 1):
                    # Generate barcode value (CC + 3-digit number)
                    barcode_value = f"CC{i:03d}"
                    
                    # Fill the number column
                    row_idx = i - start_num + 2  # Start from row 2
                    ws[f'A{row_idx}'] = barcode_value
                    
                    # Generate barcode image
                    barcode_filename = os.path.join(temp_dir, f'barcode_{i}')
                    
                    # Create barcode Code128
                    code128 = Code128(barcode_value, writer=ImageWriter())
                    code128.save(barcode_filename)
                    
                    # Add .png extension as the library saves files with this extension
                    barcode_filename_with_ext = barcode_filename + '.png'
                    
                    # Check if file exists
                    if not os.path.exists(barcode_filename_with_ext):
                        raise FileNotFoundError(f"Barcode image was not created: {barcode_filename_with_ext}")
                    
                    # Set row height
                    row_height_points = 85.04  # ~30mm in points
                    ws.row_dimensions[row_idx].height = row_height_points
                    
                    # Add image to cell
                    excel_img = ExcelImage(barcode_filename_with_ext)
                    
                    # Calculate cell dimensions
                    cell_height_pixels = int(row_height_points * 1.33)  # 1 pt ≈ 1.33 px
                    
                    # Scale image to fit cell with margins
                    target_width = max(1, cell_width_pixels - 2)
                    target_height = max(1, cell_height_pixels - 2)
                    excel_img.width = target_width
                    excel_img.height = target_height
                    
                    ws.add_image(excel_img, f'B{row_idx}')
                    
                    # Center align the image cell
                    cell = ws[f'B{row_idx}']
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Save the workbook
                wb.save(filename)
                
                messagebox.showinfo("Успех", 
                                  f"Диапазон штрихкодов от CC{start_num:03d} до CC{end_num:03d} "
                                  f"успешно сохранен в {filename}")
                
            finally:
                # Clean up temporary directory
                import shutil
                shutil.rmtree(temp_dir)
        
        except Exception as e:
            messagebox.showerror("Ошибка", f"Произошла ошибка при генерации файла:\n{str(e)}")
        finally:
            self.root.config(cursor="")


def main():
    root = Tk()
    app = BarcodeGeneratorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()