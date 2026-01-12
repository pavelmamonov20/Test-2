from openpyxl import load_workbook

# Загружаем Excel файл
wb = load_workbook('/workspace/barcodes_CC001_to_CC200.xlsx')
ws = wb.active

# Проверяем наличие изображений
images_count = len(ws._images)
print(f"Количество изображений в файле: {images_count}")

# Проверяем размеры строк
row_heights = []
for i in range(2, min(6, ws.max_row + 1)):  # Проверим первые 4 строки данных (со 2й по 5ю)
    height = ws.row_dimensions[i].height
    row_heights.append(height)
    print(f"Высота строки {i}: {height} точек")

# Проверяем значения в первой колонке
print("\nПервые 5 значений в первой колонке:")
for i in range(2, min(7, ws.max_row + 1)):
    cell_value = ws[f'A{i}'].value
    print(f"Строка {i}, колонка A: {cell_value}")

print(f"\nВсего строк в файле: {ws.max_row}")
print(f"Всего колонок в файле: {ws.max_column}")