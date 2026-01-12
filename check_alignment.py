from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Загружаем Excel файл
wb = load_workbook('/workspace/barcodes_CC001_to_CC200.xlsx')
ws = wb.active

# Проверяем выравнивание ячеек
print("Проверка выравнивания ячеек с штрих-кодами:")
for i in range(2, min(6, ws.max_row + 1)):  # Проверим первые 4 строки данных
    cell = ws[f'B{i}']
    alignment = cell.alignment
    print(f"Ячейка B{i}:")
    print(f"  horizontal: {alignment.horizontal}")
    print(f"  vertical: {alignment.vertical}")

# Также проверим размеры изображений
print(f"\nКоличество изображений в файле: {len(ws._images)}")

# Проверим первые несколько изображений
for idx, img in enumerate(ws._images[:3]):  # Проверим первые 3 изображения
    print(f"Изображение {idx+1}:")
    print(f"  Координаты: {img.anchor}")
    if hasattr(img.anchor, '_from'):
        print(f"  Позиция _from: col={img.anchor._from.col}, row={img.anchor._from.row}")
        if hasattr(img.anchor._from, 'colOff'):
            print(f"  Смещение по колонке: {img.anchor._from.colOff}")
        if hasattr(img.anchor._from, 'rowOff'):
            print(f"  Смещение по строке: {img.anchor._from.rowOff}")
    print()