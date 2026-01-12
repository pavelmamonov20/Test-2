import os
from barcode import Code128
from barcode.writer import ImageWriter
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import tempfile


def generate_barcodes_excel():
    # Создаем новую рабочую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Штрихкоды"

    # Заголовки
    ws['A1'] = 'Номер'
    ws['B1'] = 'Штрихкод'

    # Устанавливаем ширину колонок
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30

    def _excel_col_width_to_pixels(width_chars):
        # Convert Excel column width (char units) to pixels.
        return int(width_chars * 7 + 5)

    col_width = ws.column_dimensions['B'].width
    if col_width is None:
        col_width = ws.sheet_format.defaultColWidth
    if col_width is None:
        col_width = 8.43
    cell_width_pixels = int(_excel_col_width_to_pixels(col_width) * 0.98)

    # Временная директория для хранения изображений штрихкодов
    temp_dir = tempfile.mkdtemp()

    try:
        for i in range(1, 201):
            # Генерируем значение штрихкода (CC001 до CC200)
            barcode_value = f"CC{i:03d}"
            
            # Заполняем первый столбец (порядковый номер)
            ws[f'A{i+1}'] = i
            
            # Генерируем изображение штрихкода
            barcode_filename = os.path.join(temp_dir, f'barcode_{i}')
            
            # Создаем штрихкод Code128
            code128 = Code128(barcode_value, writer=ImageWriter())
            code128.save(barcode_filename)
            
            # Добавляем расширение .png, так как библиотека сохраняет файлы с этим расширением
            barcode_filename_with_ext = barcode_filename + '.png'
            
            # Проверяем, существует ли файл
            if not os.path.exists(barcode_filename_with_ext):
                raise FileNotFoundError(f"Barcode image was not created: {barcode_filename_with_ext}")
            
            # Открываем изображение и изменяем его размер для соответствия требованиям
            with PILImage.open(barcode_filename_with_ext) as img:
                original_width, original_height = img.size
            
            # Устанавливаем высоту строки в 30 мм (в Excel единицы измерения - точки)
            # 30 мм примерно равно 85.04 точкам (1 мм ≈ 2.8346 точек)
            row_height_points = 85.04
            ws.row_dimensions[i+1].height = row_height_points
            
            # Добавляем изображение в ячейку
            excel_img = Image(barcode_filename_with_ext)
            
            # Масштабируем изображение, чтобы оно поместилось в ячейку
            # Учитываем, что ширина и высота изображения должны соответствовать размеру ячейки
            cell_height_pixels = int(row_height_points * 1.33)  # 1 pt ≈ 1.33 px
            
            width_scale = cell_width_pixels / original_width
            height_scale = cell_height_pixels / original_height
            
            new_width = int(original_width * width_scale)
            new_height = int(original_height * min(height_scale, width_scale))
            
            excel_img.width = new_width
            excel_img.height = new_height
            
            ws.add_image(excel_img, f'B{i+1}')
    
        # Сохраняем файл
        wb.save('barcodes_CC001_to_CC200.xlsx')
        print("Excel-файл с штрихкодами успешно создан: barcodes_CC001_to_CC200.xlsx")
        
    finally:
        # Удаляем временные файлы
        import shutil
        shutil.rmtree(temp_dir)


if __name__ == "__main__":
    generate_barcodes_excel()
